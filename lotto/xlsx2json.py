# -*- coding: utf-8 -*-

import openpyxl
import json


def main():
    xlsx = 'excel.xlsx'
    wb = openpyxl.load_workbook(filename=xlsx)
    ws = wb['excel']

    data = {}
    data['당첨정보'] = []
    infos = data['당첨정보']
    for i in range(4, len(ws['A']) + 1):
        info = {}
        info['회차'] = ws.cell(row=i, column=2).value
        info['추첨일'] = ws.cell(row=i, column=3).value
        info['당첨번호'] = []
        numbers = info['당첨번호']
        for j in range(14, 21):
            numbers.append(ws.cell(row=i, column=j).value)
        infos.append(info)

    with open('lotto.json', 'w', encoding='utf-8') as f:
        f.write(json.dumps(data, ensure_ascii=False))


if __name__ == '__main__':
    main()
